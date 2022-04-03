"""
Taller 5
Problema 1
Bianca

Crear un archivo llamado cotizaciones.xlsx
que almacene las cotizaciones de la compañía
con las siguientes columnas:
Nombre (nombre del cliente),
cantidad,
monto.
a. Tener una función que permita leer el archivo creado.
b. Tener una función que permita leer una columna del archivo creado.

"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


# función que permita leer el archivo creado
def leer_archivo():
    try:

        dif = pd.read_excel('cotizaciones.xlsx', header=0)
        print("\n Leyendo el archivo creado\n")
        print(dif)

    except FileNotFoundError:
        print("Archivo no encontrado")

    except pd.errors.EmptyDataError:
        print("No data")

    except Exception as e:
        print("A ocurrido un error\nFavor volver a intentar ")

    return

# --------------------------------------------------------------------------------------------------------------


# Función que permita leer una columna del archivo creado.
def leer_columna():

    try:

        filesheet = 'cotizaciones.xlsx'

        wb = load_workbook(filesheet)

        column = input("\nIngrese La letra correspondienta a la columna que desea leer"
                       "\n(B Nombre), (C Cantidad), (D Monto) ")

        sheet = wb.active

        ncol = column_index_from_string(column)

        for cell, in sheet.iter_rows(min_col=ncol, max_col=ncol, values_only=True):

            print(cell)

    except Exception as e:
        print("A ocurrido un error\nFavor volver a intentar ")

    return

# --------------------------------------------------------------------------------------------------------------


data = {}
nombre = []
cantidad = []
monto = []

print("Programa de Cotizaciones")

try:

    while True:

        nom = input("Nombre del Cliente: ")
        nombre.append(nom)

        cant = int(input("Cantidad: "))
        cantidad.append(cant)

        mon = input("Monto: ")
        monto.append(mon)

        op = input("Presione X para salir ")

        if op == 'x' or op == 'X':
            break

    zipped = list(zip(nombre, cantidad, monto))

    df = pd.DataFrame(zipped, columns=['nombre', 'cantidad', 'monto'])

    # print(df)

    # Crear un archivo llamado cotizaciones.xlsx que almacene las cotizaciones de la compañía
    df.to_excel('cotizaciones.xlsx')

    leer_archivo()

    leer_columna()

except ValueError:

    print("Valor Invalido")

except Exception as e:

    print("A ocurrido un error\nFavor volver a intentar ")
